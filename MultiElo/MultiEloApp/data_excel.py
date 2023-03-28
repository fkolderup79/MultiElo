import pandas as pd
import openpyxl
from .multielo import calculate

def calc_excel(uploaded_file):

    wb = openpyxl.load_workbook(uploaded_file)

    # Get Data Frames for each Excel tabs (Settings, Distribution and Results)
    df_settings = wb["Settings"]
    df_distr = wb["Distribution"]
    df_result = wb["Results"]

    # Convert Data Frames to Data (Settings, Distribution and Results
    t_log = []
    s_settings, t_log = get_settings_from_df(df_settings, t_log)
    t_distr, t_log = get_distr_from_df(df_distr, t_log)
    t_result, t_log = get_result_from_df(df_result, s_settings, t_distr, t_log)

    t_player_class = calculate(s_settings, t_result, t_distr)

    return t_player_class, t_result


def get_settings_from_df(df_settings, t_log):
    # This function returns the Settings parameters from the Data Frame (Excel tab 'Settings')
    #--- K-Factor, E-Factor and ELO Base are required and must have a positive numeric value
    #--- Multiplier is optional and must have a positive numeric value and will be defaulted to 1 if not found
    #--- X-Distribution is optional and must have a positive numeric value and will be defaulted to 0 if not found
    t_value = []
    for row in df_settings.iter_rows(min_row=1, min_col=1, max_row=5, max_col=2):
        label = ''
        for cell in row:
            if cell.col_idx == 1:
                label = cell.value
            elif cell.col_idx == 2:
                value = ''
                if not pd.isnull(cell.value):
                    try:
                        value = int(cell.value)
                    except ValueError:
                        t_log = log_error_value_numeric(t_log, label, df_settings.title, cell)
                elif label != 'Multiplier' and label != 'X-Distribution':
                    t_log = log_error_value_initial(t_log, label, df_settings.title, cell)
                t_value.append({"Label": label, "Value": value, "Cell": cell.coordinate})
    k = ''
    e = ''
    b = ''
    m = ''
    x = ''
    for s_value in t_value:
        if s_value["Label"] == 'K-Factor':
            k = s_value["Value"]
        elif s_value["Label"] == 'E-Factor':
            e = s_value["Value"]
        elif s_value["Label"] == 'ELO Base':
            b = s_value["Value"]
        elif s_value["Label"] == 'Multiplier':
            m = s_value["Value"]
        elif s_value["Label"] == 'X-Distribution':
            x = s_value["Value"]
    if k == '':
        t_log = log_error_label_value_defined(t_log, "K-Factor", df_settings.title)
    elif k <= 0:
        t_log = log_error_label_value_positive(t_log, "K-Factor", df_settings.title)
    if e == '':
        t_log = log_error_label_value_defined(t_log, "E-Factor", df_settings.title)
    elif e <= 0:
        t_log = log_error_label_value_positive(t_log, "E-Factor", df_settings.title)
    if e == '':
        t_log = log_error_label_value_defined(t_log, "ELO Base", df_settings.title)
    elif e <= 0:
        t_log = log_error_label_value_positive(t_log, "ELO Base", df_settings.title)
    if m == '':
        m = 0
    if x == '':
        x = 0
    s_settings = {'K-factor': k, 'E-factor': e, 'ELO Base': b, 'Multiplier': m, 'X-Distribution': x}
    return s_settings, t_log


def get_distr_from_df(df_distr, t_log):
    # This function returns the Distribution data from the Data Frame (Excel tab 'Distribution')
    t_distr = []
    for row in df_distr.iter_rows(min_row=1, min_col=1, max_row=1000, max_col=3):
        found = 'X'
        for cell in row:
            if cell.col_idx == 1:
                if not pd.isnull(cell.value):
                    try:
                        count_from = int(cell.value)
                        if count_from <= 0:
                            t_log = log_error_value_positive(t_log, "Count From", df_distr.title, cell)
                    except ValueError:
                        t_log = log_error_value_numeric(t_log, "Count From", df_distr.title, cell)
                else:
                    found = ''
            elif cell.col_idx == 2:
                if not pd.isnull(cell.value):
                    try:
                        position_to = int(cell.value)
                        if position_to <= 0:
                            t_log = log_error_value_positive(t_log, "Position To", df_distr.title, cell)
                    except ValueError:
                        t_log = log_error_value_numeric(t_log, "Position To", df_distr.title, cell)
                else:
                    t_log = log_error_value_initial(t_log, "Position To", df_distr.title, cell)
                    found = ''
            elif cell.col_idx == 3:
                if not pd.isnull(cell.value):
                    try:
                        distr = int(cell.value)
                        if position_to <= 0:
                            t_log = log_error_value_positive(t_log, "Distribution", df_distr.title, cell)
                    except ValueError:
                        t_log = log_error_value_numeric(t_log, "Distribution", df_distr.title, cell)
                else:
                    t_log = log_error_value_initial(t_log, "Distribution", df_distr.title, cell)
                    found = ''
        if found == 'X':
            t_distr.append({'Count From': count_from, 'Position To': position_to, 'Distribution': distr})
        else:
            break
    t_distr.sort(reverse=False, key=lambda x: x.get('Position To'))
    t_distr.sort(reverse=False, key=lambda x: x.get('Count From'))
    return t_distr, t_log


def get_result_from_df(df_result, s_settings, t_distr, t_log):
    game_id = 0
    break_flag = ''
    t_result = []
    for row in df_result.iter_rows(min_row=2, min_col=1, max_row=100000, max_col=1000):
        game_id += 1
        t_player_id = []
        player_count = 0
        data = []
        for cell in row:
            if pd.isnull(cell.value):
                break
            if cell.col_idx == 1:
                datetime = cell.value
                try:
                    date = datetime.date()
                except AttributeError:
                    t_log = log_error_date_invalid_format(t_log, df_result.title, cell)
            elif cell.col_idx == 2:
                if not pd.isnull(cell.value) and cell.value != '':
                    distr_string = convert_distr_value(cell.value)
                    if distr_string == '':
                        t_log = log_error_distr_invalid_format(t_log, df_result.title, cell)
                else:
                    distr_string = ''
            else:
                if cell.value[0:2] == '&=':
                    value = cell.value[2:]
                else:
                    value = cell.value
                value = value_strip(value)
                if value != '':
                    if cell.value[0:2] == '&=':
                        value = '&=' + value
                    for s_player_id in t_player_id:
                        if s_player_id["Player ID"] == value:
                            t_log = log_error_duplicate_player_id(t_log, value, df_result.title, cell)
                    t_player_id.append({"Player ID": value})
                    player_count += 1
                else:
                    break
        if t_player_id == []:
            break
        if distr_string == '':
            distr_string = get_distr_string(player_count, t_distr, s_settings["X-Distribution"], t_log)
            if distr_string == '':
                t_log = log_error_game_id_distr(t_log, game_id, player_count)
        data.append(date)
        data.append(player_count)
        data.append(distr_string)
        for s_player_id in t_player_id:
            data.append(s_player_id["Player ID"])
        if t_result == []:
            t_result = {game_id: data}
        else:
            t_result[game_id] = data
    return t_result, t_log


def get_distr_string(player_count, t_distr, factor, t_log):
    t_game_distr = []
    if factor != '' and factor > 0:
        t_game_distr = calculate_distr_by_factor(player_count, factor)
    else:
        t_game_distr = convert_distr_all(player_count, t_distr)
    distr_string = convert_to_distr_string(t_game_distr)
    return distr_string


def convert_distr_value(value):
    distr_string = ''
    t_distr = []
    position = 0
    distr = ''
    for c in value:
        if c.isnumeric():
            distr = distr + c
        elif c == '-' and distr != '' and int(distr) != 0:
            position += 1
            t_distr.append({"Position": position, "Distribution": distr})
            distr = ''
        else:
            t_distr =[]
            break
    if distr != '':
        t_distr.append({"Position": position, "Distribution": distr})
    if t_distr != []:
        t_distr_adj = recalculate_distr(t_distr)
        distr_string = convert_to_distr_string(t_distr_adj)
    return distr_string


def recalculate_distr(t_distr):
    t_distr_adj = []
    distr_total = 0
    for s_distr in t_distr:
        distr_total += float(s_distr["Distribution"])
    if distr_total == 100:
        t_distr_adj = t_distr
    else:
        for s_distr in t_distr:
            distr = float(s_distr["Distribution"]) * 100 / distr_total
            t_distr_adj.append({"Position": s_distr["Position"], "Distribution": distr})
    return t_distr_adj


def calculate_distr_by_factor(player_count, factor):
    t_distr = []
    position = 1
    distr = 1
    while position < player_count:
        t_distr.append({"Position": position, "Distribution": distr})
        position += 1
        distr = distr / ((100 + factor) / 100)
    t_distr_adj = recalculate_distr(t_distr)
    return t_distr_adj


def convert_distr_all(player_count, t_distr_all):
    t_distr = []
    count_from = 0
    for s_distr_all in t_distr_all:
        if s_distr_all["Count From"] <= player_count:
            count_from = s_distr_all["Count From"]
            position_count = s_distr_all["Position To"]
        else:
            break
    position = 0
    while position_count > 0:
        position += 1
        position_count -= 1
        for s_distr_all in t_distr_all:
            if s_distr_all["Count From"] == count_from and s_distr_all["Position To"] >= position:
                t_distr.append({"Position": position, "Distribution": s_distr_all["Distribution"]})
                break
    t_distr_adj = recalculate_distr(t_distr)
    return t_distr


def convert_to_distr_string(t_distr_game):
    distr_string = ''
    for s_distr_game in t_distr_game:
        if distr_string != '':
            distr_string = distr_string + '-'
        distr_string = distr_string + str(s_distr_game["Distribution"])
    return distr_string


def value_strip(value):
    # This function will strip a value for spaces prior or after
    #--- Example value '  Player 1 ' will return value 'Player 1'
    value_out = ''
    value_part = ''
    for c in value:
        value_part = value_part + c
        if c != ' ':
            value_out = value_out + value_part
            value_part = ''
    return value_out


def log_error_value_numeric(t_log, label, sheet, cell):
    t_log.append('"' + label + '" value "' + cell.value + ' " is not numeric in Excel "' + sheet + '".' + cell.coordinate)
    return t_log


def log_error_value_initial(t_log, label, sheet, cell):
    t_log.append('"' + label +  ' "value is initial in Excel "' + sheet + '".' + cell.coordinate)
    return t_log


def log_error_value_positive(t_log, label, sheet, cell):
    t_log.append('"' + label + ' " is defined as Zero or less (not allowed) in Excel Sheet "' + sheet + '".' + cell.coordinate)
    return t_log


def log_error_label_value_defined(t_log, label, sheet):
    t_log.append('"' + label + ' " is not defined in Excel Sheet "' + sheet + '"')
    return t_log


def log_error_label_value_positive(t_log, label, sheet):
    t_log.append('"' + label + ' " is defined as Zero or less (not allowed) in Excel Sheet "' + sheet + '"')
    return t_log


def log_error_duplicate_player_id(t_log, player_id, sheet, cell):
    t_log.append('Duplicate Player ID "' + player_id + '" in Excel "' + sheet + '".' + cell.coordinate)
    return t_log


def log_error_date_invalid_format(t_log, sheet, cell):
    t_log.append('Invalid date format for value "' + cell.value + '" in Excel "' + sheet + '".' + cell.coordinate)
    return t_log


def log_error_distr_invalid_format(t_log, sheet, cell):
    t_log.append('Invalid distribution format for value "' + cell.value + '" in Excel "' + sheet + '".' + cell.coordinate)
    return t_log


def log_error_game_id_distr(t_log, game_id, player_count):
    t_log.append('Not able to determine distribution for Game ID ' + game_id + ' for Player Count ' + player_count)